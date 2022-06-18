import openpyxl
import os


def write_from_xlsx(file):
    
    def file_name(cell):
        num_type = type(sh.cell(row = i, column = 1).value)
        if num_type == float:
            num = str(int(sh.cell(row = i, column = 1).value))
        elif num_type == int:
            num = str(sh.cell(row = i, column = 1).value)
        elif num_type == str:
            num = sh.cell(row = i, column = 1).value
        return num  

    def split_title(cell):
        string1 = sh.cell(row = i, column=2).value.split(" / ")
        string2 = string1[0].replace("*", "")
        return string2

    def split_key(cell):
        return sh.cell(row = i, column=3).value.replace(" – ", " - ").split(" - ")
    
    wb = openpyxl.load_workbook(file)
    sh = wb.active
    m_row = sh.max_row

    folder_name = file.split("/")[-1].split(".")[0]
    folder_path = file.rsplit("/", 1)[0]
    folder = os.path.join(folder_path + "/", folder_name)
    os.mkdir(folder)

    for i in range(1, m_row + 1):
        
        n = file_name(i)
        txt = split_title(i)
        keys = split_key(i)
        
        txt_file = folder + "/" + n + ".txt"
        with open(txt_file, 'w+', encoding='utf-8') as tf:
            tf.write(txt)
        
        key_file = folder + "/" + n + ".key"
        with open(key_file, 'w+', encoding='utf-8') as kf:
            for key in keys:
                kf.write(key)
                kf.write('\n')